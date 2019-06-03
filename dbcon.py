import openpyxl as xl
from pymongo import MongoClient as mc
import datetime as dt
def xlsx():
    p=xl.load_workbook("Brand.xlsx")
    ob = p.active


    conn = mc('localhost',27017)
    db = conn.Medicine.THCs
    db1 = conn.Medicine.SubTHCs
    db2 = conn.Medicine.manufacturer
    db3 = conn.Medicine.MedBrands
    db4 = conn.Medicine.medicine



    for i in range(2,320798):
      on = ob.cell(row = i,column=1)
      off= ob.cell(row= i- 1,column=1)
      if on.value != off.value:
        path={"Name": on.value,"Description":" ","created":dt.datetime.utcnow(),"LastModified":dt.datetime.utcnow()}
        db.insert_one(path)


      on1 = ob.cell(row = i,column=2)
      off1= ob.cell(row= i- 1,column=2)
      if on1.value != off1.value:
        path={"Thcs": on.value,"SubThcs":on1.value,"Description":" ","created":dt.datetime.utcnow(),"LastModified":dt.datetime.utcnow()}
        db1.insert_one(path)


      on2 = ob.cell(row = i,column=7)
      off2= ob.cell(row= i- 1,column=7)

      if on2.value != off2.value:
        path={"ManufacturerName": on2.value,"Address":" ","created":dt.datetime.utcnow(),"LastModified":dt.datetime.utcnow()}
        db2.insert_one(path)


      on32 = ob.cell(row = i,column=3)
      off32= ob.cell(row= i- 1,column=3)

      on31 = ob.cell(row = i,column=6)
      off31 = ob.cell(row= i- 1,column=6)

      if on31.value != off31.value and on32.value != off32.value:
        path={"ManufactName":on2.value,"Composition":on31.value,"MedicineBrand": on32.value,"Description":" ","created":dt.datetime.utcnow(),"LastModified":dt.datetime.utcnow()}
        db3.insert_one(path)


      on41 = ob.cell(row = i,column=4)
      off41= ob.cell(row= i- 1,column=4)    


      on42 = ob.cell(row = i,column=5)
      off42 = ob.cell(row= i- 1,column=5)

      on43 = ob.cell(row = i,column=9)
      off43 = ob.cell(row= i- 1,column=9)


      on44 = ob.cell(row = i,column=8)
      off44= ob.cell(row= i- 1,column=8)
      if on41.value != off41.value and on42.value != off42.value and on43.value != off43.value and on44.value != off44.value:
        path={"SubThc":on1.value,"medForm":on42.value,"MedicinBrand":on32.value,"HsnCode":on41.value,"MaxRP":on43.value,"PrimaryPack":on44.value,"SecondaryPack":" ","TertiaryPack": " ","created":dt.datetime.utcnow(),"LastModified":dt.datetime.utcnow()}
        db4.insert_one(path)




xlsx()


